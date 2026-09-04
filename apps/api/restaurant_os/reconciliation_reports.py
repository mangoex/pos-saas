from __future__ import annotations

import io
from datetime import datetime, timedelta, timezone
from typing import Any
from zoneinfo import ZoneInfo

import openpyxl
import sqlalchemy as sa
from openpyxl.styles import Border, Font, PatternFill, Side
from sqlalchemy.orm import Session

from . import models
from .operations import (
    ORGANIZATION_ID,
    AuthorizationError,
    _id,
    authorize_branch_scope,
    require_permission,
)

UTC = timezone.utc


def _branch_day_bounds_utc(
    session: Session, branch_id: str, date_str: str
) -> tuple[datetime, datetime]:
    branch = (
        session.execute(
            sa.select(models.branches.c.timezone).where(models.branches.c.id == branch_id)
        )
        .mappings()
        .first()
    )
    tz_name = branch["timezone"] if branch and branch.get("timezone") else "America/Mazatlan"
    try:
        tz = ZoneInfo(str(tz_name))
    except Exception:
        tz = ZoneInfo("America/Mazatlan")

    dt = datetime.strptime(date_str, "%Y-%m-%d")
    local_start = datetime(dt.year, dt.month, dt.day, 0, 0, 0, 0, tzinfo=tz)
    local_end = datetime(dt.year, dt.month, dt.day, 23, 59, 59, 999999, tzinfo=tz)
    return local_start.astimezone(UTC), local_end.astimezone(UTC)


def get_branch_daily_reconciliation(
    session: Session,
    branch_id: str,
    date_str: str,
    actor_id: str | None = None,
) -> dict[str, Any]:
    """Generates complete daily cut and financial reconciliation for a branch."""
    if actor_id:
        authorize_branch_scope(session, actor_id, "dashboard.read", branch_id)

    start_utc, end_utc = _branch_day_bounds_utc(session, branch_id, date_str)

    # 1. Branch info
    branch = (
        session.execute(sa.select(models.branches).where(models.branches.c.id == branch_id))
        .mappings()
        .first()
    )
    branch_name = branch["name"] if branch else "Sucursal"

    # 2. Cash shifts
    shifts = (
        session.execute(
            sa.select(models.cash_shifts)
            .where(
                models.cash_shifts.c.branch_id == branch_id,
                models.cash_shifts.c.opened_at >= start_utc,
                models.cash_shifts.c.opened_at <= end_utc,
            )
            .order_by(models.cash_shifts.c.opened_at.desc())
        )
        .mappings()
        .all()
    )

    initial_cash_cents = sum(s["opening_cash_cents"] for s in shifts) if shifts else 0
    shift_ids = [s["id"] for s in shifts]
    closures = (
        session.execute(
            sa.select(models.cash_shift_closures).where(
                models.cash_shift_closures.c.cash_shift_id.in_(shift_ids)
            )
        )
        .mappings()
        .all()
        if shift_ids
        else []
    )

    if closures:
        physical_cash_count_cents = sum(
            int((c["summary_snapshot"] or {}).get("closing_cash_cents") or 0) for c in closures
        )
    else:
        physical_cash_count_cents = initial_cash_cents

    # 3. Orders and Payments
    payments = (
        session.execute(
            sa.select(
                models.payments.c.id,
                models.payments.c.order_id,
                models.payments.c.method,
                models.payments.c.amount_cents,
                models.payments.c.status,
                models.orders.c.folio.label("order_folio"),
                models.orders.c.owner_name.label("order_owner_name"),
                models.orders.c.customer_snapshot.label("order_customer_snapshot"),
                models.orders.c.total_cents.label("order_total_cents"),
            )
            .select_from(
                models.payments.outerjoin(
                    models.orders,
                    models.payments.c.order_id == models.orders.c.id,
                )
            )
            .where(
                models.payments.c.branch_id == branch_id,
                models.payments.c.status == "confirmed",
                models.payments.c.created_at >= start_utc,
                models.payments.c.created_at <= end_utc,
            )
        )
        .mappings()
        .all()
    )

    card_payments_cents = 0
    transfer_payments_cents = 0
    cash_sales_cents = 0
    credit_sales_cents = 0
    transfers_breakdown: list[dict[str, Any]] = []
    credit_clients_breakdown: list[dict[str, Any]] = []

    for p in payments:
        amount = p["amount_cents"]
        method = (p["method"] or "").lower()
        cust_snap = p.get("order_customer_snapshot") or {}
        cust_name = cust_snap.get("name") or p.get("order_owner_name") or "Cliente General"
        cust_phone = cust_snap.get("phone") or "—"
        folio = p.get("order_folio") or f"T-{p['id'][:6].upper()}"

        if method in ("card", "credit_card", "debit_card"):
            card_payments_cents += amount
        elif method in ("transfer", "bank_transfer", "spei"):
            transfer_payments_cents += amount
            transfers_breakdown.append(
                {
                    "ticket_folio": folio,
                    "customer_name": cust_name,
                    "customer_phone": cust_phone,
                    "amount": float(amount) / 100.0,
                }
            )
        elif method in ("credit", "customer_credit"):
            credit_sales_cents += amount
            credit_clients_breakdown.append(
                {
                    "ticket_folio": folio,
                    "customer_name": cust_name,
                    "customer_phone": cust_phone,
                    "amount": float(amount) / 100.0,
                }
            )
        else:
            cash_sales_cents += amount

    total_sales_with_tax_cents = (
        card_payments_cents + transfer_payments_cents + cash_sales_cents + credit_sales_cents
    )

    # 4. Purchases and Supplier expenses in cash
    purchases = (
        session.execute(
            sa.select(models.purchase_documents).where(
                models.purchase_documents.c.branch_id == branch_id,
                models.purchase_documents.c.status == "confirmed",
                models.purchase_documents.c.paid_from_cash.is_(True),
                models.purchase_documents.c.created_at >= start_utc,
                models.purchase_documents.c.created_at <= end_utc,
            )
        )
        .mappings()
        .all()
    )

    suppliers_breakdown: list[dict[str, Any]] = []
    supplier_expenses_cents = 0
    for idx, pur in enumerate(purchases, 1):
        sup = (
            session.execute(
                sa.select(models.suppliers).where(models.suppliers.c.id == pur["supplier_id"])
            )
            .mappings()
            .first()
        )
        sup_name = sup["commercial_name"] if sup else "Proveedor Local"
        total_mxn = float(pur["total"])
        supplier_expenses_cents += int(total_mxn * 100)
        suppliers_breakdown.append(
            {
                "no": idx,
                "provider_name": sup_name,
                "amount": total_mxn,
                "observations": f"Folio: {pur['folio']} ({pur['document_type']})",
            }
        )

    # 5. Cash movements (Gastos fijos, Retiros, Depósitos)
    movements = (
        session.execute(
            sa.select(
                models.cash_movements.c.id,
                models.cash_movements.c.movement_type,
                models.cash_movements.c.amount_cents,
                models.cash_movements.c.reason,
                models.cash_movements.c.reference,
                models.cash_movements.c.concept_snapshot,
                models.cash_movement_concept_versions.c.name.label("concept_name"),
                models.cash_movement_concepts.c.code.label("concept_code"),
            )
            .select_from(
                models.cash_movements.outerjoin(
                    models.cash_movement_concepts,
                    models.cash_movements.c.concept_id == models.cash_movement_concepts.c.id,
                ).outerjoin(
                    models.cash_movement_concept_versions,
                    (
                        models.cash_movements.c.concept_version_id
                        == models.cash_movement_concept_versions.c.id
                    ),
                )
            )
            .where(
                models.cash_movements.c.branch_id == branch_id,
                models.cash_movements.c.created_at >= start_utc,
                models.cash_movements.c.created_at <= end_utc,
            )
        )
        .mappings()
        .all()
    )

    fixed_expenses_breakdown: list[dict[str, Any]] = []
    withdrawals_breakdown: list[dict[str, Any]] = []
    fixed_expenses_cents = 0
    cash_withdrawals_cents = 0
    cash_deposits_cents = 0

    fix_idx = 1
    w_idx = 1
    for m in movements:
        mtype = m["movement_type"]
        amt = m["amount_cents"]
        cname = m["concept_name"] or (m["concept_snapshot"] or {}).get("name") or "Gasto Operativo"
        if mtype == "withdrawal":
            is_vault = (
                "retiro" in cname.lower()
                or "boveda" in cname.lower()
                or "caja fuerte" in cname.lower()
            )
            if is_vault:
                cash_withdrawals_cents += amt
                withdrawals_breakdown.append(
                    {
                        "no": w_idx,
                        "folio": f"RET-{m['id'][:6].upper()}",
                        "amount": float(amt) / 100.0,
                        "recipient_name": m["reference"] or m["reason"] or "Encargado / Bóveda",
                    }
                )
                w_idx += 1
            else:
                fixed_expenses_cents += amt
                fixed_expenses_breakdown.append(
                    {
                        "no": fix_idx,
                        "expense_type": cname,
                        "amount": float(amt) / 100.0,
                        "observations": m["reason"] or "Gasto menor de sucursal",
                    }
                )
                fix_idx += 1
        elif mtype == "deposit":
            cash_deposits_cents += amt

    # 6. Calculations (Exact Balance Formula)
    expected_cash_cents = (
        initial_cash_cents
        + cash_sales_cents
        + cash_deposits_cents
        - (supplier_expenses_cents + fixed_expenses_cents + cash_withdrawals_cents)
    )
    difference_cents = physical_cash_count_cents - expected_cash_cents

    # 7. Persistent Audit record lookup
    audit_row = (
        session.execute(
            sa.select(models.reconciliation_audit_logs).where(
                models.reconciliation_audit_logs.c.branch_id == branch_id,
                models.reconciliation_audit_logs.c.date == date_str,
            )
        )
        .mappings()
        .first()
    )

    if audit_row:
        audit = {
            "reviewed": bool(audit_row["reviewed"]),
            "audited_by_user_id": audit_row["audited_by_user_id"],
            "audited_at": audit_row["audited_at"].isoformat() if audit_row["audited_at"] else None,
            "notes": audit_row["notes"],
        }
    else:
        audit = {
            "reviewed": False,
            "audited_by_user_id": None,
            "audited_at": None,
            "notes": None,
        }

    return {
        "branch_id": branch_id,
        "branch_name": branch_name,
        "date": date_str,
        "balance": {
            "initial_cash": float(initial_cash_cents) / 100.0,
            "total_sales_with_tax": float(total_sales_with_tax_cents) / 100.0,
            "card_payments": float(card_payments_cents) / 100.0,
            "transfer_payments": float(transfer_payments_cents) / 100.0,
            "credit_sales": float(credit_sales_cents) / 100.0,
            "cash_sales": float(cash_sales_cents) / 100.0,
            "supplier_expenses": float(supplier_expenses_cents) / 100.0,
            "fixed_expenses": float(fixed_expenses_cents) / 100.0,
            "cash_withdrawals": float(cash_withdrawals_cents) / 100.0,
            "cash_deposits": float(cash_deposits_cents) / 100.0,
            "expected_cash_in_register": float(expected_cash_cents) / 100.0,
            "physical_cash_count": float(physical_cash_count_cents) / 100.0,
            "difference": float(difference_cents) / 100.0,
        },
        "suppliers_breakdown": suppliers_breakdown,
        "fixed_expenses_breakdown": fixed_expenses_breakdown,
        "transfers_breakdown": transfers_breakdown,
        "credit_clients_breakdown": credit_clients_breakdown,
        "withdrawals_breakdown": withdrawals_breakdown,
        "audit": audit,
    }


def get_multi_branch_consolidated_report(
    session: Session,
    date_from_str: str,
    date_to_str: str,
    branch_id: str | None = None,
    actor_id: str | None = None,
) -> dict[str, Any]:
    """Aggregates daily reconciliations across all or selected branches."""
    if actor_id:
        if branch_id:
            authorize_branch_scope(session, actor_id, "dashboard.read", branch_id)
        else:
            require_permission(session, actor_id, "dashboard.read", None)

    branches_query = sa.select(models.branches)
    if branch_id:
        branches_query = branches_query.where(models.branches.c.id == branch_id)
    branches = session.execute(branches_query).mappings().all()

    supplier_totals: dict[str, float] = {}
    fixed_expense_totals: dict[str, float] = {}
    branch_summaries: list[dict[str, Any]] = []

    total_sales = 0.0
    total_cards = 0.0
    total_transfers = 0.0
    total_credits = 0.0
    total_suppliers = 0.0
    total_fixed = 0.0
    total_withdrawals = 0.0
    total_expected = 0.0

    # Parse date range
    dt_from = datetime.strptime(date_from_str, "%Y-%m-%d")
    dt_to = datetime.strptime(date_to_str, "%Y-%m-%d")
    curr = dt_from
    days = []
    while curr <= dt_to:
        days.append(curr.strftime("%Y-%m-%d"))
        curr += timedelta(days=1)

    for b in branches:
        b_id = b["id"]
        b_name = b["name"]
        b_sales = 0.0
        b_expenses = 0.0
        for day in days:
            rep = get_branch_daily_reconciliation(session, b_id, day, actor_id)
            bal = rep["balance"]
            b_sales += bal["total_sales_with_tax"]
            b_expenses += bal["supplier_expenses"] + bal["fixed_expenses"]
            total_sales += bal["total_sales_with_tax"]
            total_cards += bal["card_payments"]
            total_transfers += bal["transfer_payments"]
            total_credits += bal["credit_sales"]
            total_suppliers += bal["supplier_expenses"]
            total_fixed += bal["fixed_expenses"]
            total_withdrawals += bal["cash_withdrawals"]
            total_expected += bal["expected_cash_in_register"]

            for sup in rep["suppliers_breakdown"]:
                sname = sup["provider_name"]
                supplier_totals[sname] = supplier_totals.get(sname, 0.0) + sup["amount"]

            for fexp in rep["fixed_expenses_breakdown"]:
                ename = fexp["expense_type"]
                fixed_expense_totals[ename] = fixed_expense_totals.get(ename, 0.0) + fexp["amount"]

        branch_summaries.append(
            {
                "branch_id": b_id,
                "branch_name": b_name,
                "total_sales": b_sales,
                "total_expenses": b_expenses,
            }
        )

    return {
        "date_from": date_from_str,
        "date_to": date_to_str,
        "branches": branch_summaries,
        "supplier_totals": supplier_totals,
        "fixed_expense_totals": fixed_expense_totals,
        "summary": {
            "total_sales": total_sales,
            "total_cards": total_cards,
            "total_transfers": total_transfers,
            "total_credits": total_credits,
            "total_suppliers": total_suppliers,
            "total_fixed": total_fixed,
            "total_withdrawals": total_withdrawals,
            "total_expected_cash": total_expected,
        },
    }


def update_reconciliation_audit_status(
    session: Session,
    branch_id: str,
    date_str: str,
    reviewed: bool,
    notes: str | None = None,
    auditor_id: str | None = None,
) -> dict[str, Any]:
    """Records persistent auditor validation state in database."""
    if auditor_id:
        try:
            authorize_branch_scope(session, auditor_id, "audit.read", branch_id)
        except AuthorizationError:
            try:
                authorize_branch_scope(session, auditor_id, "branch.admin.access", branch_id)
            except AuthorizationError:
                require_permission(session, auditor_id, "admin.manage", branch_id)

    now = datetime.now(timezone.utc)
    existing = (
        session.execute(
            sa.select(models.reconciliation_audit_logs).where(
                models.reconciliation_audit_logs.c.branch_id == branch_id,
                models.reconciliation_audit_logs.c.date == date_str,
            )
        )
        .mappings()
        .first()
    )

    if existing:
        session.execute(
            models.reconciliation_audit_logs.update()
            .where(models.reconciliation_audit_logs.c.id == existing["id"])
            .values(
                reviewed=bool(reviewed),
                audited_by_user_id=auditor_id or "admin-user",
                notes=notes or "",
                audited_at=now,
                updated_at=now,
            )
        )
    else:
        session.execute(
            models.reconciliation_audit_logs.insert().values(
                id=_id(),
                organization_id=ORGANIZATION_ID,
                branch_id=branch_id,
                date=date_str,
                reviewed=bool(reviewed),
                audited_by_user_id=auditor_id or "admin-user",
                notes=notes or "",
                audited_at=now,
                created_at=now,
                updated_at=now,
            )
        )
    session.commit()

    return {
        "branch_id": branch_id,
        "date": date_str,
        "reviewed": bool(reviewed),
        "audited_by_user_id": auditor_id or "admin-user",
        "audited_at": now.isoformat(),
        "notes": notes or "",
    }


def export_reconciliation_workbook(
    session: Session,
    branch_id: str,
    month: int,
    year: int,
    actor_id: str | None = None,
) -> io.BytesIO:
    """Generates standard Excel (.xlsx) matching the structure of Kiwi Multi-Branch Cuts."""
    if actor_id:
        authorize_branch_scope(session, actor_id, "dashboard.read", branch_id)

    wb = openpyxl.Workbook()
    # Sheet 1: Resumen
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    ws_resumen["A1"] = f"BALANCE CONSOLIDADO MENSUAL ({month:02d}/{year})"
    ws_resumen["A1"].font = Font(name="Calibri", size=14, bold=True)

    headers = [
        "Concepto",
        "Monto Total ($)",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_resumen.cell(3, col_idx, h)
        cell.font = header_font
        cell.fill = header_fill

    # Calculate month total
    date_from = f"{year}-{month:02d}-01"
    next_m = month + 1 if month < 12 else 1
    next_y = year if month < 12 else year + 1
    last_day = (datetime(next_y, next_m, 1) - timedelta(days=1)).day
    date_to = f"{year}-{month:02d}-{last_day:02d}"

    rep = get_multi_branch_consolidated_report(session, date_from, date_to, branch_id, actor_id)
    summary = rep["summary"]

    rows = [
        ("Ventas Totales con Impuestos", summary["total_sales"]),
        ("(-) Pagos con Tarjeta", summary["total_cards"]),
        ("(-) Ingresos por Transferencias", summary["total_transfers"]),
        ("(-) Clientes a Crédito", summary["total_credits"]),
        ("(-) Pago a Proveedores en Efectivo", summary["total_suppliers"]),
        ("(-) Gastos Fijos en Efectivo", summary["total_fixed"]),
        ("(-) Retiros en Efectivo a Bóveda", summary["total_withdrawals"]),
        ("(=) Efectivo Neto Esperado", summary["total_expected_cash"]),
    ]

    for r_idx, (lbl, val) in enumerate(rows, 4):
        ws_resumen.cell(r_idx, 1, lbl).border = border
        c = ws_resumen.cell(r_idx, 2, val)
        c.border = border
        c.number_format = '"$"#,##0.00'

    # Sheet 2: Master
    ws_master = wb.create_sheet(title="Master")
    ws_master["A1"] = "PLANTILLA DE CORTE DIARIO CONSOLIDADO"
    ws_master["A1"].font = Font(name="Calibri", size=12, bold=True)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
